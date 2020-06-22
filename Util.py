import openpyxl
import glob

class Utils:
    def __init__(self):
        self.ext_txt = ".txt"
        self.ext_dat = ".dat"
        self.ext_xlsx = ".xlsx"

    """
    get file lists in target dir by target ext
    :param
        path : target dir + "*." + target ext
    :return
        ['target dir/file_name.target ext', 'target dir/file_name.target ext' ...]
    """
    def get_files_from_dir(self, path):
        return glob.glob(path)

    def read_txt(self, path):
        result_list = []
        with open(path, "r") as f:
            while True:
                tmp_line = f.readline().replace("\n", "")
                if tmp_line == "":
                    break

                result_list.append(tmp_line)

        return result_list

    def make_excel(self, path, data_dict):
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        row = 1
        sheet.cell(row=row, column=1, value="index")
        sheet.cell(row=row, column=2, value='file_name')
        sheet.cell(row=row, column=3, value='contents')
        for fn_key, read_line_arr in data_dict.items():
            row += 1
            sheet.cell(row=row, column=1, value=str(row - 1))
            sheet.cell(row=row, column=2, value=fn_key)
            col = 3
            for read_line in read_line_arr:
                sheet.cell(row=row, column=col, value=float(read_line))
                col += 1

        workbook.save(filename=path + self.ext_xlsx)
